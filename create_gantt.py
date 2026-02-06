from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Project Schedule"

# Define colors
header_blue = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
light_blue = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
phase_blue = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
task_gray = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
task_dark = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
white_font = Font(color="FFFFFF", bold=True)
black_font = Font(color="000000")
bold_font = Font(bold=True)
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# Project info
ws['A1'] = "JobIntel Platform"
ws['A1'].font = Font(size=18, italic=True, color="4472C4")
ws.merge_cells('A1:C1')

ws['R1'] = "Project Schedule"
ws['R1'].font = Font(size=18, bold=True)
ws.merge_cells('R1:V1')

# Start date
ws['A3'] = "Start Week"
ws['A3'].fill = phase_blue
ws['A3'].font = white_font
ws['B3'] = "Oct 5, 2025"
ws['B3'].fill = light_blue
ws.merge_cells('B3:C3')

# Define weeks (36 weeks from Oct 2025 to Jun 2026)
start_date = datetime(2025, 10, 5)
weeks = []
for i in range(36):
    week_start = start_date + timedelta(weeks=i)
    weeks.append(week_start)

# Week headers
ws['A5'] = "Week"
ws['A5'].fill = header_blue
ws['A5'].font = white_font
ws['A5'].alignment = Alignment(horizontal='center')

ws['A6'] = "Starting"
ws['A6'].fill = header_blue
ws['A6'].font = white_font
ws['A6'].alignment = Alignment(horizontal='center')

# Add week numbers and dates
for i, week in enumerate(weeks):
    col = i + 2
    cell = ws.cell(row=5, column=col)
    cell.value = i + 1
    cell.fill = header_blue
    cell.font = white_font
    cell.alignment = Alignment(horizontal='center')
    
    date_cell = ws.cell(row=6, column=col)
    date_cell.value = week.strftime("%b\n%d")
    date_cell.fill = header_blue
    date_cell.font = white_font
    date_cell.alignment = Alignment(horizontal='center', wrap_text=True)

# Notes column
notes_col = len(weeks) + 2
ws.cell(row=5, column=notes_col).value = "Notes"
ws.cell(row=5, column=notes_col).fill = header_blue
ws.cell(row=5, column=notes_col).font = white_font
ws.merge_cells(start_row=5, start_column=notes_col, end_row=6, end_column=notes_col)

# Define tasks with phases, task name, start week, duration in weeks
tasks = [
    # Phase, Task, Start Week (1-based), Duration (weeks)
    ("Phase One\nResearch", "Topic Research & Selection", 1, 2),
    ("", "Business Requirements", 2, 4),
    ("", "Tools Installation", 2, 1),
    ("", "Technology Study", 3, 4),
    
    ("Phase Two\nAnalysis & Design", "Use Case Analysis", 7, 4),
    ("", "Use Case Scenarios", 7, 4),
    ("", "DFD", 7, 4),
    ("", "UI Wireframes (Figma)", 7, 5),
    ("", "Sequence Diagrams", 7, 5),
    ("", "ERD Design", 9, 3),
    ("", "Database Schema", 11, 2),
    ("", "Class Diagrams", 7, 5),
    
    ("Phase Three\nImplementation", "Frontend Implementation", 12, 7),
    ("", "Backend Setup & Database", 18, 3),
    ("", "API Development", 21, 11),
    ("", "AI Model Development", 12, 18),
    ("", "AI API Integration", 29, 3),
    
    ("Phase Four\nTesting", "Frontend Testing", 31, 2),
    ("", "Backend Testing", 31, 2),
    ("", "AI Testing", 31, 2),
    ("", "Integration Testing", 32, 2),
    
    ("Phase Five\nDocumentation", "Technical Documentation", 13, 22),
    ("", "Final Report", 33, 3),
    ("", "Presentation Preparation", 35, 2),
]

# Write tasks
current_row = 7
current_phase = ""
phase_start_row = 7

for task in tasks:
    phase, task_name, start_week, duration = task
    
    # Phase column
    if phase:
        if current_phase and phase != current_phase:
            # Merge previous phase cells
            if current_row - phase_start_row > 1:
                ws.merge_cells(start_row=phase_start_row, start_column=1, 
                              end_row=current_row-1, end_column=1)
        current_phase = phase
        phase_start_row = current_row
        ws.cell(row=current_row, column=1).value = phase
        ws.cell(row=current_row, column=1).fill = phase_blue
        ws.cell(row=current_row, column=1).font = white_font
        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal='center', 
                                                                  vertical='center',
                                                                  wrap_text=True)
    
    # Task name in column after week 1 position (we'll use merged cells)
    # For simplicity, put task name starting at column 2
    task_cell = ws.cell(row=current_row, column=2)
    
    # Draw the Gantt bar
    for w in range(start_week, start_week + duration):
        if w <= len(weeks):
            bar_cell = ws.cell(row=current_row, column=w + 1)
            bar_cell.fill = task_gray
            bar_cell.border = thin_border
    
    # Put task name in the first cell of the bar
    task_cell = ws.cell(row=current_row, column=start_week + 1)
    task_cell.value = task_name
    task_cell.font = Font(color="FFFFFF", size=9)
    task_cell.alignment = Alignment(horizontal='left', vertical='center')
    
    current_row += 1

# Merge last phase
if current_row - phase_start_row > 1:
    ws.merge_cells(start_row=phase_start_row, start_column=1, 
                  end_row=current_row-1, end_column=1)

# Add "PROJECT END" in notes column
end_row = 7
for i in range(current_row - 7):
    cell = ws.cell(row=end_row + i, column=notes_col)
    cell.border = thin_border

# Write PROJECT END vertically
project_end = "PROJECT END"
for i, char in enumerate(project_end):
    if end_row + i < current_row:
        ws.cell(row=end_row + i, column=notes_col).value = char
        ws.cell(row=end_row + i, column=notes_col).alignment = Alignment(horizontal='center')

# Set column widths
ws.column_dimensions['A'].width = 12
for i in range(2, notes_col):
    ws.column_dimensions[get_column_letter(i)].width = 4
ws.column_dimensions[get_column_letter(notes_col)].width = 8

# Set row heights
for row in range(5, current_row):
    ws.row_dimensions[row].height = 20

# Add borders to all cells in the chart area
for row in range(5, current_row):
    for col in range(1, notes_col + 1):
        cell = ws.cell(row=row, column=col)
        if not cell.fill or cell.fill.start_color.index == '00000000':
            cell.border = thin_border

# Save
wb.save('Docs/ProjectSchedule_Gantt.xlsx')
print("Gantt chart created successfully: Docs/ProjectSchedule_Gantt.xlsx")
